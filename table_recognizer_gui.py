from __future__ import annotations

import base64
import json
import os
import queue
import re
import threading
import time
import urllib.error
import urllib.request
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk

import cv2
import numpy as np

try:
    from preprocess_frame import PreprocessFrame
except Exception:  # pragma: no cover - module import must not break the GUI
    PreprocessFrame = None  # type: ignore[assignment,misc]

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    DND_FILES = None
    TkinterDnD = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先运行 install_and_run.bat") from exc


DEFAULT_API = "http://192.168.10.26:8087/table-recognition"
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
# Preprocessed images produced by the "图片预处理" tab, keyed by resolved path.
# When present, recognition uses this JPEG instead of the raw photo.
PREPROCESSED_IMAGES: dict[str, bytes] = {}
HEADER_WORDS = {
    "序号", "名称", "姓名", "编号", "代码", "规格", "型号", "单位", "数量",
    "单价", "金额", "日期", "时间", "备注", "项目", "类别", "部门", "地址",
    "电话", "税率", "合计", "内容", "状态", "说明", "产品", "物料", "订单",
    "name", "code", "id", "date", "amount", "price", "quantity", "total",
    "description", "remark", "unit", "item", "type", "status",
}
ARCHIVE_TABLE_HEADERS = [
    "日期",
    "档号",
    "案卷名称（工程名称）",
    "文书/采购/合同整理（整理）",
    "工程/科技档案整理（整理）",
    "装订（卷）",
    "编制页面（页/卷）",
    "条目录入（条）",
    "盖档号章（个）",
    "打印卷内目录；案卷目录",
    "文档扫描（页数）",
    "图纸扫描（页数）",
    "挂接条目（条）",
    "装盒（盒）",
]


@dataclass
class ParsedTable:
    rows: list[list[str]]
    merges: list[tuple[int, int, int, int]] = field(default_factory=list)
    header_rows: int = 0
    source: str = ""


@dataclass
class TableGroup:
    title: str
    tables: list[ParsedTable] = field(default_factory=list)


class TableHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[tuple[str, int, int]]]] = []
        self._depth = 0
        self._rows: list[list[tuple[str, int, int]]] = []
        self._row: list[tuple[str, int, int]] | None = None
        self._cell_parts: list[str] | None = None
        self._rowspan = 1
        self._colspan = 1

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._depth += 1
            if self._depth == 1:
                self._rows = []
        elif self._depth == 1 and tag == "tr":
            self._row = []
        elif self._depth == 1 and tag in {"td", "th"}:
            values = dict(attrs)
            self._rowspan = max(1, int(values.get("rowspan") or 1))
            self._colspan = max(1, int(values.get("colspan") or 1))
            self._cell_parts = []
        elif self._cell_parts is not None and tag in {"br", "p", "div"}:
            self._cell_parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if self._depth == 1 and tag in {"td", "th"} and self._cell_parts is not None:
            text = re.sub(r"[ \t\r\f\v]+", " ", "".join(self._cell_parts))
            text = re.sub(r"\n\s*\n+", "\n", text).strip()
            if self._row is not None:
                self._row.append((text, self._rowspan, self._colspan))
            self._cell_parts = None
        elif self._depth == 1 and tag == "tr" and self._row is not None:
            self._rows.append(self._row)
            self._row = None
        elif tag == "table" and self._depth:
            if self._depth == 1:
                self.tables.append(self._rows)
            self._depth -= 1


def expand_html_table(raw_rows: list[list[tuple[str, int, int]]]) -> ParsedTable:
    rows: list[list[str]] = []
    merges: list[tuple[int, int, int, int]] = []
    pending: dict[int, int] = {}

    for row_index, raw_row in enumerate(raw_rows, start=1):
        row: list[str] = []
        col = 1

        def consume_pending() -> None:
            nonlocal col
            while col in pending:
                while len(row) < col:
                    row.append("")
                pending[col] -= 1
                if pending[col] <= 0:
                    del pending[col]
                col += 1

        for text, rowspan, colspan in raw_row:
            consume_pending()
            start_col = col
            while len(row) < start_col - 1:
                row.append("")
            row.append(text)
            for _ in range(colspan - 1):
                row.append("")
            if rowspan > 1 or colspan > 1:
                merges.append(
                    (row_index, start_col, row_index + rowspan - 1, start_col + colspan - 1)
                )
            if rowspan > 1:
                for occupied_col in range(start_col, start_col + colspan):
                    pending[occupied_col] = max(pending.get(occupied_col, 0), rowspan - 1)
            col += colspan
        consume_pending()
        rows.append(row)

    width = max((len(row) for row in rows), default=0)
    for row in rows:
        row.extend([""] * (width - len(row)))
    return ParsedTable(rows=rows, merges=merges)


def parse_html_tables(html: str) -> list[ParsedTable]:
    parser = TableHTMLParser()
    parser.feed(html or "")
    return [expand_html_table(raw) for raw in parser.tables if raw]


def is_number_like(value: str) -> bool:
    compact = re.sub(r"[\s,，￥¥$%％()（）./:\-]", "", value)
    return bool(compact) and bool(re.fullmatch(r"[0-9]+", compact))


def natural_path_key(path: Path) -> tuple:
    """Sort filenames case-insensitively while comparing digit runs numerically."""
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", path.name)
    )


def detect_header_rows(rows: list[list[str]]) -> int:
    """Return the number of leading rows that look like a title/header block."""
    candidates: list[tuple[float, int]] = []
    for index, row in enumerate(rows[:6]):
        values = [str(value).strip() for value in row if str(value).strip()]
        if len(values) < 2:
            continue
        lowered = [value.casefold() for value in values]
        keyword_hits = sum(
            1 for value in lowered if any(word == value or word in value for word in HEADER_WORDS)
        )
        numeric_ratio = sum(is_number_like(value) for value in values) / len(values)
        text_ratio = sum(any(ch.isalpha() for ch in value) for value in values) / len(values)
        long_ratio = sum(len(value) > 32 for value in values) / len(values)
        score = keyword_hits * 2.5 + text_ratio * 2.0 - numeric_ratio * 2.0 - long_ratio
        if len(values) >= 3:
            score += 0.5
        candidates.append((score, index))
    if not candidates:
        return 0
    score, index = max(candidates)
    return index + 1 if score >= 1.4 else 0


def normalize_known_header(row: list[str]) -> list[str]:
    """Repair the fixed 14-column archive header when OCR joins adjacent labels."""
    if len(row) != len(ARCHIVE_TABLE_HEADERS):
        return row
    joined = re.sub(r"\s+", "", "".join(str(value) for value in row))
    signatures = ("档号", "装订", "编制页", "条目", "打印卷", "文档扫", "图纸扫", "挂接", "装盒")
    if sum(signature in joined for signature in signatures) < 5:
        return row
    return list(ARCHIVE_TABLE_HEADERS)


def _order_quad(points: np.ndarray) -> np.ndarray:
    points = points.astype(np.float32).reshape(4, 2)
    ordered = np.zeros((4, 2), dtype=np.float32)
    sums = points.sum(axis=1)
    differences = np.diff(points, axis=1).reshape(-1)
    ordered[0] = points[np.argmin(sums)]
    ordered[2] = points[np.argmax(sums)]
    ordered[1] = points[np.argmin(differences)]
    ordered[3] = points[np.argmax(differences)]
    return ordered


def rectify_table_image(path: Path) -> bytes:
    """Detect the outer grid and perspective-warp a photographed table."""
    encoded = np.fromfile(str(path), dtype=np.uint8)
    image = cv2.imdecode(encoded, cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"无法读取图片：{path.name}")

    original_height, original_width = image.shape[:2]
    detection_scale = min(1.0, 2200.0 / max(original_height, original_width))
    small = cv2.resize(
        image,
        None,
        fx=detection_scale,
        fy=detection_scale,
        interpolation=cv2.INTER_AREA,
    )
    gray = cv2.cvtColor(small, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 51, 15
    )
    height, width = binary.shape
    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (max(35, width // 24), 1)),
    )
    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(35, height // 24))),
    )
    grid = cv2.bitwise_or(horizontal, vertical)
    grid = cv2.morphologyEx(
        grid,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (9, 9)),
        iterations=2,
    )
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    minimum_area = width * height * 0.20
    quad: np.ndarray | None = None
    sorted_contours = sorted(contours, key=cv2.contourArea, reverse=True)
    for contour in sorted_contours:
        if cv2.contourArea(contour) < minimum_area:
            break
        hull = cv2.convexHull(contour)
        perimeter = cv2.arcLength(hull, True)
        approximation = cv2.approxPolyDP(hull, 0.02 * perimeter, True)
        if len(approximation) == 4 and cv2.isContourConvex(approximation):
            quad = approximation.reshape(4, 2).astype(np.float32) / detection_scale
            break
    if quad is None and sorted_contours and cv2.contourArea(sorted_contours[0]) >= minimum_area:
        box = cv2.boxPoints(cv2.minAreaRect(sorted_contours[0]))
        quad = box.astype(np.float32) / detection_scale

    if quad is not None:
        center = quad.mean(axis=0)
        quad = center + (quad - center) * 1.05
        quad[:, 0] = np.clip(quad[:, 0], 0, original_width - 1)
        quad[:, 1] = np.clip(quad[:, 1], 0, original_height - 1)
        source = _order_quad(quad)
        target_width = int(max(np.linalg.norm(source[1] - source[0]), np.linalg.norm(source[2] - source[3])))
        target_height = int(max(np.linalg.norm(source[3] - source[0]), np.linalg.norm(source[2] - source[1])))
        if target_width >= 400 and target_height >= 400:
            destination = np.array(
                [[0, 0], [target_width - 1, 0], [target_width - 1, target_height - 1], [0, target_height - 1]],
                dtype=np.float32,
            )
            transform = cv2.getPerspectiveTransform(source, destination)
            image = cv2.warpPerspective(
                image,
                transform,
                (target_width, target_height),
                flags=cv2.INTER_CUBIC,
                borderMode=cv2.BORDER_CONSTANT,
                borderValue=(255, 255, 255),
            )
            image = cv2.copyMakeBorder(
                image, 12, 12, 12, 12, cv2.BORDER_CONSTANT, value=(255, 255, 255)
            )

    ok, output = cv2.imencode(".jpg", image, [cv2.IMWRITE_JPEG_QUALITY, 95])
    if not ok:
        raise ValueError(f"图片预处理失败：{path.name}")
    return output.tobytes()


def enhance_low_resolution_image(image_bytes: bytes) -> bytes | None:
    """Return a conservative 2x OCR supplement for a low-resolution page."""
    image = cv2.imdecode(np.frombuffer(image_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    if image is None or max(image.shape[:2]) >= 1800:
        return None
    scale = min(2.0, 2000.0 / max(image.shape[:2]))
    enlarged = cv2.resize(
        image,
        None,
        fx=scale,
        fy=scale,
        interpolation=cv2.INTER_CUBIC,
    )
    ok, output = cv2.imencode(".jpg", enlarged, [cv2.IMWRITE_JPEG_QUALITY, 96])
    return output.tobytes() if ok else None


def ocr_needs_low_resolution_supplement(image_bytes: bytes, payload: dict) -> bool:
    """Judge effective text resolution instead of relying on image size alone."""
    image = cv2.imdecode(np.frombuffer(image_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    if image is None or max(image.shape[:2]) >= 1800:
        return False

    heights: list[float] = []
    scores: list[float] = []
    for result in payload.get("result", {}).get("tableRecResults", []):
        ocr = result.get("overall_ocr_res", {})
        boxes = ocr.get("rec_boxes") or []
        texts = ocr.get("rec_texts") or []
        rec_scores = ocr.get("rec_scores") or [1.0] * len(texts)
        for box, text, score in zip(boxes, texts, rec_scores):
            if not str(text).strip() or len(box) < 4:
                continue
            x1, y1, x2, y2 = map(float, box[:4])
            height = y2 - y1
            if height <= 0:
                continue
            heights.append(height)
            scores.append(float(score))

    # If the first pass found no text on a small image, a larger retry is the
    # only useful fallback. Otherwise use actual glyph size and confidence.
    if not heights:
        return True
    median_height = float(np.median(heights))
    median_score = float(np.median(scores)) if scores else 1.0
    low_confidence_ratio = sum(score < 0.70 for score in scores) / max(1, len(scores))
    return (
        median_height < 18.0
        or (median_height < 24.0 and median_score < 0.78)
        or (median_height < 24.0 and low_confidence_ratio >= 0.20)
    )


def _cluster_positions(indices: np.ndarray, gap: int) -> list[int]:
    if not len(indices):
        return []
    groups = np.split(indices, np.where(np.diff(indices) > gap)[0] + 1)
    return [int(round(float(np.median(group)))) for group in groups if len(group)]


def _merge_fragments(items: list[tuple[float, float, float, str]]) -> str:
    """Join OCR fragments that belong to one physical cell."""
    items = sorted(items, key=lambda item: (item[1], item[0]))
    lines: list[list[tuple[float, float, float, str]]] = []
    for item in items:
        if not lines or abs(item[1] - np.mean([part[1] for part in lines[-1]])) > item[2] * 0.65:
            lines.append([item])
        else:
            lines[-1].append(item)

    rendered_lines: list[str] = []
    for line in lines:
        line.sort(key=lambda item: item[0])
        combined = ""
        previous_right: float | None = None
        for left, _, height, text in line:
            text = text.strip()
            if not text:
                continue
            if not combined:
                combined = text
            else:
                overlap = 0
                limit = min(len(combined), len(text), 16)
                for size in range(limit, 0, -1):
                    if combined[-size:].casefold() == text[:size].casefold():
                        overlap = size
                        break
                code_like = bool(re.fullmatch(r"[A-Za-z0-9_./\-]+", combined + text))
                close = previous_right is not None and left - previous_right < height * 1.5
                separator = "" if overlap or combined.endswith("-") or text.startswith("-") or (code_like and close) else " "
                combined += separator + text[overlap:]
            previous_right = max(previous_right or left, left + max(height, len(text) * height * 0.4))
        if combined:
            rendered_lines.append(combined)
    return "\n".join(rendered_lines)


def rebuild_grid_table(image_bytes: bytes, ocr: dict) -> ParsedTable | None:
    """Rebuild a bordered table from real grid lines and OCR coordinates."""
    image = cv2.imdecode(np.frombuffer(image_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    boxes = ocr.get("rec_boxes") or []
    texts = ocr.get("rec_texts") or []
    scores = ocr.get("rec_scores") or [1.0] * len(texts)
    if image is None or len(boxes) < 2 or len(boxes) != len(texts):
        return None

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 51, 12
    )
    height, width = binary.shape
    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(35, height // 18))),
    )
    strong_x = np.where(vertical.sum(axis=0) > 255 * height * 0.12)[0]
    x_lines = _cluster_positions(strong_x, max(3, width // 100))
    if len(x_lines) < 3:
        return None
    # Ignore line fragments outside the main grid span.
    x_lines = [x for x in x_lines if width * 0.01 <= x <= width * 0.99]
    column_count = len(x_lines) - 1

    entries: list[dict] = []
    for box, text, score in zip(boxes, texts, scores):
        if float(score) < 0.35 or not str(text).strip() or len(box) < 4:
            continue
        x1, y1, x2, y2 = map(float, box[:4])
        recognized_text = str(text).strip()
        glyph_ratio = (x2 - x1) / max(1.0, y2 - y1)
        # A genuine Chinese glyph is roughly square. On compressed pages the
        # model may label a very narrow printed "1" plus grid noise as 一/二/I.
        if recognized_text in {"一", "二", "I", "l", "|", "丨"} and glyph_ratio <= 0.70:
            recognized_text = "1"
        center_x = (x1 + x2) / 2
        center_y = (y1 + y2) / 2
        column = int(np.searchsorted(x_lines, center_x, side="right") - 1)
        if not (0 <= column < column_count):
            continue
        entries.append(
            {
                "x1": x1,
                "x2": x2,
                "cy": center_y,
                "height": max(1.0, y2 - y1),
                "column": column,
                "text": recognized_text,
            }
        )
    if len(entries) < 2:
        return None

    median_height = float(np.median([entry["height"] for entry in entries]))
    row_tolerance = max(12.0, median_height * 0.65)

    def cluster_by_y(items: list[dict]) -> list[list[dict]]:
        clusters: list[list[dict]] = []
        for item in sorted(items, key=lambda candidate: candidate["cy"]):
            if not clusters:
                clusters.append([item])
                continue
            center = float(np.mean([part["cy"] for part in clusters[-1]]))
            if abs(item["cy"] - center) <= row_tolerance:
                clusters[-1].append(item)
            else:
                clusters.append([item])
        return clusters

    # The column with the most long text (normally an item code/name column) is
    # the most reliable row anchor. Other columns may print text at different
    # vertical positions inside the same cell and must not define row grouping.
    anchor_candidates: list[tuple[float, int, list[list[dict]]]] = []
    for column in range(column_count):
        column_items = [entry for entry in entries if entry["column"] == column]
        clusters = cluster_by_y(column_items)
        if len(clusters) < 2:
            continue
        average_length = np.mean(
            [sum(len(part["text"]) for part in cluster) for cluster in clusters]
        )
        anchor_candidates.append((len(clusters) * (1.0 + average_length / 8.0), column, clusters))
    if not anchor_candidates:
        return None
    _, anchor_column, anchor_clusters = max(anchor_candidates, key=lambda candidate: candidate[0])
    anchor_centers = [float(np.mean([item["cy"] for item in cluster])) for cluster in anchor_clusters]
    if len(anchor_centers) < 2:
        return None

    def cluster_is_data(cluster: list[dict]) -> bool:
        return any(is_number_like(str(item["text"])) for item in cluster)

    # If the anchor's first OCR cluster is textual and the following cluster is
    # numeric/code-like, the first cluster is the physical header row. This is
    # more reliable than a global Y cutoff because photographed tables often
    # slope substantially from left to right.
    anchor_data_start = next(
        (index for index, cluster in enumerate(anchor_clusters) if cluster_is_data(cluster)),
        0,
    )
    has_leading_header = anchor_data_start > 0
    data_row_centers = anchor_centers[anchor_data_start:]
    if len(data_row_centers) < 2:
        return None
    row_pitch = float(np.median(np.diff(data_row_centers)))

    header_entries: list[dict] = []
    data_entries: list[dict] = []
    column_clusters_by_column: dict[int, list[list[dict]]] = {}
    for column in range(column_count):
        clusters = cluster_by_y([entry for entry in entries if entry["column"] == column])
        if has_leading_header:
            first_data = next(
                (index for index, cluster in enumerate(clusters) if cluster_is_data(cluster)),
                len(clusters),
            )
            for cluster in clusters[:first_data]:
                header_entries.extend(cluster)
            clusters = clusters[first_data:]
        column_clusters_by_column[column] = clusters
        for cluster in clusters:
            data_entries.extend(cluster)

    # Estimate the photographed row-baseline slope from columns populated on
    # most rows. Sparse columns then use the same fitted offset instead of being
    # shifted one or two rows down on strongly skewed pages.
    offset_samples: list[tuple[float, float]] = []
    dense_series: list[tuple[float, list[float]]] = []
    dense_minimum = max(3, int(len(data_row_centers) * 0.75))
    for column, clusters in column_clusters_by_column.items():
        if len(clusters) != len(data_row_centers) or len(clusters) < dense_minimum:
            continue
        centers = [float(np.mean([item["cy"] for item in cluster])) for cluster in clusters]
        column_x = (x_lines[column] + x_lines[column + 1]) / 2
        dense_series.append((column_x, centers))
        offset_samples.append((column_x, float(np.median(np.asarray(centers) - data_row_centers))))
    if len(offset_samples) >= 2:
        slope, intercept = np.polyfit(
            [sample[0] for sample in offset_samples],
            [sample[1] for sample in offset_samples],
            1,
        )
    elif offset_samples:
        slope, intercept = 0.0, offset_samples[0][1]
    else:
        slope, intercept = 0.0, 0.0

    # Paper curvature makes that slope stronger near the top than near the
    # bottom. Fit an offset for every row using dense columns, with the global
    # fit as a fallback.
    row_offset_models: list[tuple[float, float]] = []
    for row_index, anchor_center in enumerate(data_row_centers):
        if len(dense_series) >= 2:
            row_slope, row_intercept = np.polyfit(
                [series[0] for series in dense_series],
                [series[1][row_index] - anchor_center for series in dense_series],
                1,
            )
            row_offset_models.append((float(row_slope), float(row_intercept)))
        else:
            row_offset_models.append((float(slope), float(intercept)))

    row_clusters: list[list[dict]] = [[] for _ in data_row_centers]
    for column in range(column_count):
        column_clusters = column_clusters_by_column[column]
        if not column_clusters:
            continue
        column_centers = [float(np.mean([item["cy"] for item in cluster])) for cluster in column_clusters]
        if len(column_clusters) == len(data_row_centers):
            for row_index, cluster in enumerate(column_clusters):
                row_clusters[row_index].extend(cluster)
            continue
        column_x = (x_lines[column] + x_lines[column + 1]) / 2
        for center, cluster in zip(column_centers, column_clusters):
            expected_centers = np.asarray(
                [
                    anchor_center + row_slope * column_x + row_intercept
                    for anchor_center, (row_slope, row_intercept)
                    in zip(data_row_centers, row_offset_models)
                ]
            )
            row_index = int(np.argmin(np.abs(expected_centers - center)))
            row_clusters[row_index].extend(cluster)

    if has_leading_header:
        row_clusters.insert(0, header_entries)

    rows: list[list[str]] = []
    for cluster in row_clusters:
        row = [""] * column_count
        for column in range(column_count):
            fragments = [
                (entry["x1"], entry["cy"], entry["height"], entry["text"])
                for entry in cluster
                if entry["column"] == column
            ]
            if fragments:
                row[column] = _merge_fragments(fragments)
        rows.append(row)

    if has_leading_header and rows:
        rows[0] = normalize_known_header(rows[0])

    # Chinese OCR commonly confuses tiny 1/2 glyphs with 一/二/I/l/|. Correct
    # only exact, isolated glyphs in columns whose populated values are already
    # overwhelmingly numeric, leaving real Chinese text untouched.
    numeric_confusions = {
        "一": "1",
        "I": "1",
        "l": "1",
        "|": "1",
        "丨": "1",
        "二": "2",
    }
    data_start = 1 if has_leading_header else 0
    for column in range(column_count):
        populated = [
            str(row[column]).strip()
            for row in rows[data_start:]
            if column < len(row) and str(row[column]).strip()
        ]
        if not populated:
            continue
        numeric_count = sum(is_number_like(value) for value in populated)
        compatible_count = sum(
            is_number_like(value) or value in numeric_confusions for value in populated
        )
        if numeric_count < 2 or compatible_count / len(populated) < 0.8:
            continue
        for row in rows[data_start:]:
            if column >= len(row):
                continue
            value = str(row[column]).strip()
            if value in numeric_confusions:
                row[column] = numeric_confusions[value]

    # This archive form has two mandatory flag columns three and four cells to
    # the right of the file-number column; every ordinary record contains 1 in
    # both. Low-resolution pages often lose these thinnest glyphs completely.
    code_column_counts = []
    for column in range(column_count):
        count = sum(
            bool(re.match(r"^800217-", str(row[column]).strip()))
            for row in rows[data_start:]
            if column < len(row)
        )
        code_column_counts.append(count)
    if code_column_counts and max(code_column_counts) >= 3 and 14 <= column_count <= 16:
        code_column = int(np.argmax(code_column_counts))
        mandatory_columns = (code_column + 3, code_column + 4)
        for row in rows[data_start:]:
            if code_column >= len(row) or not re.match(r"^800217-", str(row[code_column]).strip()):
                continue
            for column in mandatory_columns:
                if column < len(row) and not str(row[column]).strip():
                    row[column] = "1"

    # Repair a one-character OCR slip in a strongly repeated numeric code
    # prefix (for example 80027 -> 800217) without altering the variable suffix.
    for column in range(column_count):
        prefixes = []
        for row in rows:
            value = row[column]
            match = re.match(r"^(\d{4,})-", value)
            if match:
                prefixes.append(match.group(1))
        if len(prefixes) < 3:
            continue
        common, common_count = Counter(prefixes).most_common(1)[0]
        if common_count < max(3, len(prefixes) * 0.45):
            continue

        def one_edit_away(value: str, reference: str) -> bool:
            if abs(len(value) - len(reference)) > 1:
                return False
            if len(value) == len(reference):
                return sum(a != b for a, b in zip(value, reference)) <= 1
            shorter, longer = (value, reference) if len(value) < len(reference) else (reference, value)
            return any(longer[:index] + longer[index + 1:] == shorter for index in range(len(longer)))

        for row in rows:
            match = re.match(r"^(\d{4,})(-.+)$", row[column])
            if match and match.group(1) != common and one_edit_away(match.group(1), common):
                row[column] = common + match.group(2)

    # Recover horizontal cell merges by checking whether each internal vertical
    # boundary actually exists through the corresponding OCR row band.
    merges: list[tuple[int, int, int, int]] = []
    merge_row_centers = list(data_row_centers)
    if has_leading_header:
        merge_row_centers.insert(0, float(np.mean([entry["cy"] for entry in header_entries])))
    for row_index, center_y in enumerate(merge_row_centers, start=1):
        previous_center = merge_row_centers[row_index - 2] if row_index > 1 else center_y - median_height
        next_center = merge_row_centers[row_index] if row_index < len(merge_row_centers) else center_y + median_height
        top = int(max(0, (previous_center + center_y) / 2 + 4))
        bottom = int(min(height, (center_y + next_center) / 2 - 4))
        span_start = 1
        for boundary_column, x in enumerate(x_lines[1:-1], start=1):
            strip = vertical[top:bottom, max(0, x - 3): min(width, x + 4)]
            boundary_exists = strip.size and np.count_nonzero(strip) / strip.size >= 0.03
            if boundary_exists:
                if boundary_column > span_start:
                    merges.append((row_index, span_start, row_index, boundary_column))
                span_start = boundary_column + 1
        if column_count > span_start:
            merges.append((row_index, span_start, row_index, column_count))

    # Curved paper can move a vertical line several pixels between the top and
    # bottom of a row. False merges are much more damaging than leaving a real
    # merged header as separate cells, so coordinate-rebuilt tables keep the
    # detected column grid without speculative merges.
    table = ParsedTable(rows=rows, merges=[])
    table.header_rows = detect_header_rows(rows)
    return table


def _request_table_payload(image_bytes: bytes, api_url: str, timeout: int) -> dict:
    body = json.dumps(
        {
            "file": base64.b64encode(image_bytes).decode("ascii"),
            "fileType": 1,
            "visualize": False,
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        api_url,
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    try:
        with opener.open(request, timeout=timeout) as response:
            payload = json.load(response)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise RuntimeError(f"服务返回 HTTP {exc.code}: {detail}") from exc
    except (urllib.error.URLError, TimeoutError) as exc:
        raise RuntimeError(f"无法连接识别服务：{exc}") from exc

    if payload.get("errorCode") not in (None, 0):
        raise RuntimeError(payload.get("errorMsg") or "识别服务返回错误")
    return payload


def _tables_from_payload(payload: dict, image_bytes: bytes, source_name: str) -> list[ParsedTable]:
    result_items = payload.get("result", {}).get("tableRecResults", [])
    tables: list[ParsedTable] = []
    for result in result_items:
        grid_table = rebuild_grid_table(image_bytes, result.get("overall_ocr_res", {}))
        if grid_table is not None:
            grid_table.source = source_name
            tables.append(grid_table)
            continue
        for table_result in result.get("table_res_list", []):
            parsed = parse_html_tables(table_result.get("pred_html", ""))
            for table in parsed:
                table.source = source_name
                table.header_rows = detect_header_rows(table.rows)
                tables.append(table)
    return tables


def _code_column(table: ParsedTable) -> tuple[int, dict[str, list[str]]] | None:
    if not table.rows:
        return None
    width = max(len(row) for row in table.rows)
    candidates: list[tuple[int, int, dict[str, list[str]]]] = []
    for column in range(width):
        code_rows: dict[str, list[str]] = {}
        for row in table.rows:
            if column >= len(row):
                continue
            value = str(row[column]).strip()
            if re.match(r"^800217-", value):
                code_rows[value] = row
        candidates.append((len(code_rows), column, code_rows))
    count, column, code_rows = max(candidates, default=(0, 0, {}), key=lambda item: item[0])
    return (column, code_rows) if count >= 3 else None


def _merge_numeric_supplement(primary: ParsedTable, supplement: ParsedTable) -> None:
    """Fill primary OCR blanks with numeric values from a low-res 2x pass."""
    primary_codes = _code_column(primary)
    supplement_codes = _code_column(supplement)
    if primary_codes is None or supplement_codes is None:
        return
    primary_column, primary_rows = primary_codes
    supplement_column, supplement_rows = supplement_codes
    for code, primary_row in primary_rows.items():
        supplement_row = supplement_rows.get(code)
        if supplement_row is None:
            continue
        for relative_column in range(3, 13):
            primary_index = primary_column + relative_column
            supplement_index = supplement_column + relative_column
            if primary_index >= len(primary_row) or supplement_index >= len(supplement_row):
                continue
            if str(primary_row[primary_index]).strip():
                continue
            value = str(supplement_row[supplement_index]).strip()
            if is_number_like(value):
                primary_row[primary_index] = value


def _path_key(path: Path) -> str:
    return str(Path(path).resolve()).casefold()


def set_preprocessed(path: Path, image_bytes: bytes) -> None:
    """Store a preprocessed image (from the preprocessing tab) for a file."""
    PREPROCESSED_IMAGES[_path_key(path)] = image_bytes


def clear_preprocessed(path: Path) -> None:
    PREPROCESSED_IMAGES.pop(_path_key(path), None)


def get_preprocessed(path: Path) -> bytes | None:
    return PREPROCESSED_IMAGES.get(_path_key(path))


def load_image_bytes(path: Path) -> bytes:
    """Return the preprocessed image for ``path``, else rectify the raw photo."""
    cached = get_preprocessed(path)
    if cached is not None:
        return cached
    return rectify_table_image(path)


def recognize_image(path: Path, api_url: str, timeout: int = 180) -> list[ParsedTable]:
    image_bytes = load_image_bytes(path)
    payload = _request_table_payload(image_bytes, api_url, timeout)
    tables = _tables_from_payload(payload, image_bytes, path.name)

    enhanced_bytes = (
        enhance_low_resolution_image(image_bytes)
        if ocr_needs_low_resolution_supplement(image_bytes, payload)
        else None
    )
    if enhanced_bytes is not None and tables:
        try:
            enhanced_payload = _request_table_payload(enhanced_bytes, api_url, timeout)
            enhanced_tables = _tables_from_payload(enhanced_payload, enhanced_bytes, path.name)
        except RuntimeError:
            enhanced_tables = []
        for table in tables:
            table_codes = _code_column(table)
            if table_codes is None:
                continue
            primary_code_set = set(table_codes[1])
            best = max(
                enhanced_tables,
                key=lambda candidate: len(
                    primary_code_set & set((_code_column(candidate) or (0, {}))[1])
                ),
                default=None,
            )
            if best is not None:
                _merge_numeric_supplement(table, best)
    return tables


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name).strip(" '") or "表格"
    base = base[:31]
    candidate = base
    number = 2
    while candidate.casefold() in used:
        suffix = f"_{number}"
        candidate = base[: 31 - len(suffix)] + suffix
        number += 1
    used.add(candidate.casefold())
    return candidate


def next_available_path(path: Path) -> Path:
    if not path.exists():
        return path
    number = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{number}{path.suffix}")
        if not candidate.exists():
            return candidate
        number += 1


def excel_cell_value(value: object, force_text: bool = False) -> object:
    """Convert safe OCR numerals to real Excel numbers without damaging IDs."""
    if force_text or not isinstance(value, str):
        return value
    text = value.strip()
    if not text or "\n" in text:
        return value
    normalized = text.replace(",", "").replace("，", "")
    integer_match = re.fullmatch(r"([+-]?)(\d+)", normalized)
    if integer_match:
        digits = integer_match.group(2)
        # Excel only preserves 15 significant digits, and leading zeroes are
        # normally meaningful in codes rather than arithmetic values.
        if len(digits) > 15 or (len(digits) > 1 and digits.startswith("0")):
            return text
        return int(normalized)
    decimal_match = re.fullmatch(r"([+-]?)(\d+)\.(\d+)", normalized)
    if decimal_match:
        integer_part = decimal_match.group(2)
        significant_digits = len((integer_part + decimal_match.group(3)).lstrip("0"))
        if (len(integer_part) > 1 and integer_part.startswith("0")) or significant_digits > 15:
            return text
        return float(normalized)
    return value


def _is_cjk_char(char: str) -> bool:
    """Whether a single character is a CJK ideograph (basic + Extension A)."""
    if not char:
        return False
    code = ord(char)
    return 0x4E00 <= code <= 0x9FFF or 0x3400 <= code <= 0x4DBF


def _is_identifier_label(label: str, identifier_words: tuple[str, ...]) -> bool:
    """Tell whether a header label denotes an ID/text column.

    Identifier words (日期/档号/编号/代码/邮编/电话/id) mean "keep this column as
    text".  A naive substring test wrongly flags count columns whose header merely
    contains an ID word, e.g. ``盖档号章（个）`` contains ``档号``.  Treat an
    identifier word as a real token only when it is NOT embedded between two CJK
    characters — i.e. it appears at a label edge or next to a separator, as in
    ``档号``, ``日期`` or ``合同编号``/``产品代码``.
    """
    for word in identifier_words:
        start = 0
        while True:
            index = label.find(word, start)
            if index == -1:
                break
            before = label[index - 1] if index > 0 else ""
            after = label[index + len(word)] if index + len(word) < len(label) else ""
            embedded_between_cjk = _is_cjk_char(before) and _is_cjk_char(after)
            if not embedded_between_cjk:
                return True
            start = index + 1
    return False


def write_groups_to_xlsx(groups: list[TableGroup], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "PaddleOCR 表格识别工具"
    used_names: set[str] = set()
    thin = Side(style="thin", color="B7C3D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for group_number, group in enumerate(groups, start=1):
        title = safe_sheet_name(group.title or f"表格{group_number}", used_names)
        sheet = workbook.create_sheet(title)
        row_offset = 0
        first_header_rows = group.tables[0].header_rows if group.tables else 0
        all_merges: list[tuple[int, int, int, int]] = []
        force_text_columns: set[int] = set()
        if first_header_rows and group.tables and group.tables[0].rows:
            header_width = max(len(row) for row in group.tables[0].rows[:first_header_rows])
            identifier_words = ("日期", "时间", "档号", "编号", "代码", "邮编", "电话", "id")
            for column in range(header_width):
                label = "".join(
                    str(row[column]) if column < len(row) else ""
                    for row in group.tables[0].rows[:first_header_rows]
                ).casefold()
                if _is_identifier_label(label, identifier_words):
                    force_text_columns.add(column + 1)

        for table in group.tables:
            for row_number, row in enumerate(table.rows, start=1):
                for col_number, value in enumerate(row, start=1):
                    is_header = row_offset == 0 and row_number <= first_header_rows
                    output_value = value if is_header else excel_cell_value(
                        value, force_text=col_number in force_text_columns
                    )
                    cell = sheet.cell(row=row_offset + row_number, column=col_number, value=output_value)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = border
                    if is_header:
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
            for r1, c1, r2, c2 in table.merges:
                all_merges.append((r1 + row_offset, c1, r2 + row_offset, c2))
            row_offset += len(table.rows)

        for r1, c1, r2, c2 in all_merges:
            if r1 != r2 or c1 != c2:
                try:
                    sheet.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                except ValueError:
                    pass

        if first_header_rows:
            sheet.freeze_panes = f"A{first_header_rows + 1}"
        sheet.sheet_view.showGridLines = False
        for column in range(1, sheet.max_column + 1):
            max_length = 0
            for row in range(1, min(sheet.max_row, 500) + 1):
                value = sheet.cell(row=row, column=column).value
                if value:
                    max_length = max(max_length, max(len(part) for part in str(value).splitlines()))
            sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 10), 45)

    if not workbook.sheetnames:
        workbook.create_sheet("无识别结果")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


class TableRecognizerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("PaddleOCR 表格识别工具")
        self.root.geometry("920x680")
        self.root.minsize(760, 560)
        self.files: list[Path] = []
        self.events: queue.Queue[tuple[str, object]] = queue.Queue()
        self.output_dir = tk.StringVar(value=str(Path.home() / "Desktop"))
        self.api_url = tk.StringVar(value=DEFAULT_API)
        self.output_mode = tk.StringVar(value="combined")
        self.status = tk.StringVar(value="拖入图片，或点击“添加图片”")
        self._build_ui()
        self.root.after(100, self._drain_events)

    def _build_ui(self) -> None:
        # Two-tab layout: recognition list + image preprocessing.
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        list_tab = ttk.Frame(self.notebook, padding=14)
        self.notebook.add(list_tab, text="待识别列表")
        self._build_list_tab(list_tab)

        preprocess_tab = ttk.Frame(self.notebook)
        self.notebook.add(preprocess_tab, text="图片预处理")
        self.preprocess_frame = PreprocessFrame(preprocess_tab, self.files, self)
        self.preprocess_frame.pack(fill="both", expand=True)
        self._preprocess_frame = self.preprocess_frame

    def _build_list_tab(self, outer: ttk.Frame) -> None:
        ttk.Label(outer, text="待识别图片", font=("Microsoft YaHei UI", 13, "bold")).pack(anchor="w")
        list_frame = ttk.Frame(outer)
        list_frame.pack(fill="both", expand=True, pady=(8, 8))
        self.listbox = tk.Listbox(list_frame, selectmode="extended", font=("Microsoft YaHei UI", 10))
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=scrollbar.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        if DND_FILES:
            self.listbox.drop_target_register(DND_FILES)
            self.listbox.dnd_bind("<<Drop>>", self._on_drop)

        buttons = ttk.Frame(outer)
        buttons.pack(fill="x", pady=(0, 10))
        ttk.Button(buttons, text="添加图片", command=self._choose_files).pack(side="left")
        ttk.Button(buttons, text="移除选中", command=self._remove_selected).pack(side="left", padx=6)
        ttk.Button(buttons, text="清空", command=self._clear_files).pack(side="left")
        ttk.Button(buttons, text="上移", command=lambda: self._move(-1)).pack(side="right", padx=3)
        ttk.Button(buttons, text="下移", command=lambda: self._move(1)).pack(side="right", padx=3)

        options = ttk.LabelFrame(outer, text="输出设置", padding=10)
        options.pack(fill="x")
        ttk.Radiobutton(
            options, text="合并为一个 XLSX（自动续表，每张新表一个 Sheet）",
            variable=self.output_mode, value="combined"
        ).grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Radiobutton(
            options, text="每张图片生成一个 XLSX（图片内多表分别放入 Sheet）",
            variable=self.output_mode, value="separate"
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(3, 8))
        ttk.Label(options, text="保存文件夹：").grid(row=2, column=0, sticky="w")
        ttk.Entry(options, textvariable=self.output_dir).grid(row=2, column=1, sticky="ew", padx=6)
        ttk.Button(options, text="选择…", command=self._choose_output_dir).grid(row=2, column=2)
        ttk.Label(options, text="API 地址：").grid(row=3, column=0, sticky="w", pady=(7, 0))
        ttk.Entry(options, textvariable=self.api_url).grid(row=3, column=1, columnspan=2, sticky="ew", padx=6, pady=(7, 0))
        options.columnconfigure(1, weight=1)

        action = ttk.Frame(outer)
        action.pack(fill="x", pady=(12, 6))
        self.start_button = ttk.Button(action, text="开始识别并生成 XLSX", command=self._start)
        self.start_button.pack(side="right")
        self.progress = ttk.Progressbar(action, mode="determinate")
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, 12))

        ttk.Label(outer, textvariable=self.status).pack(anchor="w")
        self.log = tk.Text(outer, height=8, state="disabled", font=("Consolas", 9))
        self.log.pack(fill="x", pady=(5, 0))
        if not DND_FILES:
            self._append_log("提示：未安装 tkinterdnd2，拖放不可用；可运行 install_and_run.bat 安装。")

    def _on_drop(self, event: object) -> None:
        paths = self.root.tk.splitlist(getattr(event, "data", ""))
        self._add_paths([Path(path) for path in paths])

    def _choose_files(self) -> None:
        paths = filedialog.askopenfilenames(
            title="选择表格图片",
            filetypes=[("图片文件", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.webp"), ("所有文件", "*.*")],
        )
        self._add_paths([Path(path) for path in paths])

    def _add_paths(self, paths: list[Path]) -> None:
        known = {str(path).casefold() for path in self.files}
        for path in paths:
            if path.is_dir():
                children = sorted(p for p in path.iterdir() if p.suffix.casefold() in IMAGE_SUFFIXES)
                self._add_paths(children)
            elif path.suffix.casefold() in IMAGE_SUFFIXES and str(path).casefold() not in known:
                self.files.append(path)
                known.add(str(path).casefold())
        self.files.sort(key=natural_path_key)
        self._refresh_list()

    def _refresh_list(self) -> None:
        self.listbox.delete(0, "end")
        for index, path in enumerate(self.files, start=1):
            self.listbox.insert("end", f"{index:02d}. {path.name}")
        self.status.set(f"已添加 {len(self.files)} 张图片")
        self._notify_preprocess_files()

    def _notify_preprocess_files(self) -> None:
        frame = getattr(self, "_preprocess_frame", None)
        if frame is not None:
            frame.files_changed()

    def _remove_selected(self) -> None:
        selected = set(self.listbox.curselection())
        kept = [path for index, path in enumerate(self.files) if index not in selected]
        self.files[:] = kept
        self._refresh_list()

    def _clear_files(self) -> None:
        self.files.clear()
        self._refresh_list()

    def _move(self, direction: int) -> None:
        selected = list(self.listbox.curselection())
        if len(selected) != 1:
            return
        old = selected[0]
        new = old + direction
        if 0 <= new < len(self.files):
            self.files[old], self.files[new] = self.files[new], self.files[old]
            self._refresh_list()
            self.listbox.selection_set(new)

    def _choose_output_dir(self) -> None:
        path = filedialog.askdirectory(initialdir=self.output_dir.get() or None)
        if path:
            self.output_dir.set(path)

    def _append_log(self, text: str) -> None:
        self.log.configure(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    # --- Preprocessing cache bridge (called by the preprocessing tab) ------ #
    def set_preprocessed(self, path: Path, image_bytes: bytes) -> None:
        set_preprocessed(path, image_bytes)

    def clear_preprocessed(self, path: Path) -> None:
        clear_preprocessed(path)

    def log_status(self, text: str) -> None:
        self._append_log(text)
        self.status.set(text)

    def _start(self) -> None:
        if not self.files:
            messagebox.showwarning("没有图片", "请先拖入或选择图片。")
            return
        output_dir = Path(self.output_dir.get().strip())
        if not self.output_dir.get().strip():
            messagebox.showwarning("没有输出目录", "请选择 XLSX 保存文件夹。")
            return
        self.start_button.configure(state="disabled")
        self.progress.configure(maximum=len(self.files), value=0)
        self.status.set("正在识别…")
        self._append_log("=" * 60)
        worker = threading.Thread(
            target=self._worker,
            args=(list(self.files), output_dir, self.api_url.get().strip(), self.output_mode.get()),
            daemon=True,
        )
        worker.start()

    def _worker(self, files: list[Path], output_dir: Path, api_url: str, mode: str) -> None:
        try:
            files = sorted(files, key=natural_path_key)
            output_dir.mkdir(parents=True, exist_ok=True)
            recognized: list[tuple[Path, list[ParsedTable]]] = []
            for index, path in enumerate(files, start=1):
                started = time.perf_counter()
                self.events.put(("log", f"[{index}/{len(files)}] 识别：{path.name}"))
                tables = recognize_image(path, api_url)
                elapsed = time.perf_counter() - started
                self.events.put(("log", f"    找到 {len(tables)} 个表格，耗时 {elapsed:.2f} 秒"))
                recognized.append((path, tables))
                self.events.put(("progress", index))

            outputs: list[Path] = []
            if mode == "separate":
                for path, tables in recognized:
                    groups = [
                        TableGroup(title=f"表格{number}", tables=[table])
                        for number, table in enumerate(tables, start=1)
                    ]
                    output = next_available_path(output_dir / f"{path.stem}.xlsx")
                    write_groups_to_xlsx(groups, output)
                    outputs.append(output)
            else:
                groups: list[TableGroup] = []
                for path, tables in recognized:
                    for table_index, table in enumerate(tables):
                        # Multiple tables detected in one image are always separate sheets.
                        starts_new = not groups or table_index > 0 or table.header_rows > 0
                        if starts_new:
                            title = path.stem if table_index == 0 else f"{path.stem}_表{table_index + 1}"
                            groups.append(TableGroup(title=title, tables=[table]))
                            reason = "首张表" if len(groups) == 1 else "检测到表头/同图多表"
                            self.events.put(("log", f"    新建 Sheet：{title}（{reason}）"))
                        else:
                            groups[-1].tables.append(table)
                            self.events.put(("log", f"    无表头，追加到 Sheet：{groups[-1].title}"))
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output = next_available_path(output_dir / f"表格识别_{stamp}.xlsx")
                write_groups_to_xlsx(groups, output)
                outputs.append(output)

            self.events.put(("done", outputs))
        except Exception as exc:
            self.events.put(("error", f"{type(exc).__name__}: {exc}"))

    def _drain_events(self) -> None:
        try:
            while True:
                kind, value = self.events.get_nowait()
                if kind == "log":
                    self._append_log(str(value))
                elif kind == "progress":
                    self.progress.configure(value=int(value))
                elif kind == "done":
                    outputs = list(value)  # type: ignore[arg-type]
                    self.start_button.configure(state="normal")
                    self.status.set(f"完成：生成 {len(outputs)} 个 XLSX 文件")
                    for path in outputs:
                        self._append_log(f"已保存：{path}")
                    messagebox.showinfo("识别完成", "已生成：\n" + "\n".join(str(path) for path in outputs))
                elif kind == "error":
                    self.start_button.configure(state="normal")
                    self.status.set("处理失败")
                    self._append_log(str(value))
                    messagebox.showerror("处理失败", str(value))
        except queue.Empty:
            pass
        self.root.after(100, self._drain_events)


def main() -> None:
    root = TkinterDnD.Tk() if TkinterDnD else tk.Tk()
    try:
        ttk.Style(root).theme_use("vista")
    except tk.TclError:
        pass
    TableRecognizerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
